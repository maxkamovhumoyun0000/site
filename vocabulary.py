import io
import random
import re
from typing import List, Dict, Any, Tuple
from openpyxl import load_workbook
from auth import normalize_level_to_cefr
from db import get_conn
from openpyxl import Workbook
import io

EXPECTED_VOCAB_HEADERS = ['Level', 'Word', 'translation_uz', 'translation_ru', 'Example Sentence 1', 'Example Sentence 2']


def _normalize_word_key(value: Any) -> str:
    return str(value or '').strip().lower()


def get_available_vocabulary_levels(user_level: str) -> list:
    """User levelidan past va teng leveldagi darajalarni qaytaradi"""
    level_order = ['A1', 'A2', 'B1', 'B2', 'C1']
    raw = normalize_level_to_cefr(user_level)
    if raw not in level_order:
        try:
            idx = level_order.index((user_level or "").strip().upper())
            return level_order[: idx + 1]
        except ValueError:
            return level_order
    try:
        idx = level_order.index(raw)
        return level_order[: idx + 1]
    except ValueError:
        return level_order


def vocab_language_for_subject(subject: str) -> str:
    """words.language ustuni bilan mos: English → en, Russian → ru (import/AI bilan bir xil)."""
    subj = (subject or "").strip().lower()
    if "russian" in subj:
        return "ru"
    return "en"


def parse_excel_bytes(file_bytes: bytes) -> tuple[List[Dict[str, Any]], list[str]]:
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True)
    sheet = wb.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return [], []

    def normalize_header(v: Any) -> str:
        s = str(v).strip().lower() if v is not None else ''
        s = s.replace('-', '_')
        s = re.sub(r'\s+', '_', s)
        # Keep only letters/numbers/underscore so variants still match (e.g. "translation ru")
        s = re.sub(r'[^a-z0-9_]', '', s)
        return s

    raw_headers = [normalize_header(h) for h in rows[0]]
    results = []
    for row in rows[1:]:
        if all(c is None for c in row):
            continue
        entry = {}
        example_parts = []

        for i, cell in enumerate(row):
            if i >= len(raw_headers):
                continue
            key = raw_headers[i]
            if key.startswith('example'):
                if cell:
                    example_parts.append(str(cell))
            else:
                mapped = {
                    'level': 'level',
                    'word': 'word',
                    'translation_uz': 'translation_uz',
                    'translation_ru': 'translation_ru',
                    'definition': 'definition'
                }.get(key, key)
                entry[mapped] = cell

        entry['example'] = '\n'.join(example_parts) if example_parts else ''
        results.append(entry)

    return results, raw_headers


def _validate_excel_headers(raw_headers: list[str], language: str) -> None:
    """
    Validate that the uploaded sheet has the mandatory columns.
    We keep this strict to make teacher mistakes visible immediately.
    """
    required = {'level', 'word', 'translation_uz', 'definition'}
    if language == 'en':
        # English template: translation_ru + mandatory definition.
        required.add('translation_ru')
    elif language == 'ru':
        # Russian template: definition (plus optional example sentences).
        pass
    else:
        raise ValueError(f"Unsupported language for import: {language}")

    missing = sorted([c for c in required if c not in raw_headers])
    if missing:
        raise ValueError(
            "Incorrect Excel headers. Missing: "
            + ", ".join(missing)
            + ". Please upload the provided template."
        )


def import_words_from_excel(file_bytes: bytes, file_name: str, added_by: int | None, subject: str, language: str):
    parsed, raw_headers = parse_excel_bytes(file_bytes)
    subject = (subject or '').strip().title()
    language = (language or '').strip().lower()
    _validate_excel_headers(raw_headers, language)
    conn = get_conn()
    cur = conn.cursor()

    cur.execute(
        '''
        INSERT INTO vocabulary_imports (file_name, added_by, subject, language)
        VALUES (%s, %s, %s, %s)
        RETURNING id
        ''',
        (file_name, added_by, subject, language),
    )
    row = cur.fetchone()
    import_id = row["id"] if row else None

    inserted = 0
    skipped_existing = 0
    skipped_invalid = 0
    duplicate_words = []

    for entry in parsed:
        word = (entry.get('word') or '').strip()
        normalized_word = str(word or '').strip()
        if not _normalize_word_key(normalized_word):
            skipped_invalid += 1
            continue

        cur.execute("""
            SELECT id FROM words 
            WHERE subject = ? 
              AND language = ? 
              AND LOWER(TRIM(word)) = LOWER(TRIM(?))
            LIMIT 1
        """, (subject, language, normalized_word))

        if cur.fetchone():
            skipped_existing += 1
            duplicate_words.append(normalized_word)
            continue

        level = entry.get('level') or ''
        translation_uz = entry.get('translation_uz') or ''
        translation_ru = entry.get('translation_ru') or ''
        definition = entry.get('definition') or ''
        example = entry.get('example') or ''
        if not str(definition).strip():
            skipped_invalid += 1
            continue

        cur.execute('''
            INSERT INTO words 
            (word, subject, language, level, translation_uz, translation_ru, definition, example, added_by)
            VALUES (?,?,?,?,?,?,?,?,?)
        ''', (normalized_word, subject, language, level, translation_uz, translation_ru, definition, example, added_by))

        inserted += 1

    conn.commit()
    conn.close()

    skipped = skipped_existing + skipped_invalid

    if inserted == 0 and skipped == len(parsed):
        # If every row was skipped due to duplicates or empty words, tell the teacher.
        # (We don't know exact reason, but it is most often wrong headers / empty 'Word' column.)
        pass

    return {
        'import_id': import_id,
        'inserted': inserted,
        'skipped': skipped,
        'skipped_existing': skipped_existing,
        'skipped_invalid': skipped_invalid,
        'duplicates': duplicate_words,
        'total': len(parsed)
    }


def search_words(subject: str, query: str, levels: List[str] | None = None) -> List[Dict[str, Any]]:
    if not subject:
        return []

    vocab_lang = vocab_language_for_subject(subject)

    conn = get_conn()
    try:
        cur = conn.cursor()
        base_sql = 'SELECT * FROM words WHERE LOWER(subject)=LOWER(?) AND LOWER(language)=LOWER(?)'
        params: list[Any] = [subject.strip(), vocab_lang]
        level_clause = ''
        if levels:
            placeholders = ",".join(["?"] * len(levels))
            level_clause = f" AND level IN ({placeholders})"
            params.extend(levels)
        filter_sql = ""
        q_raw = query.strip().lower()
        if q_raw:
            q = f"%{q_raw}%"
            filter_sql = (
                " AND (LOWER(word) LIKE ? OR LOWER(translation_uz) LIKE ? OR LOWER(translation_ru) LIKE ? "
                "OR LOWER(definition) LIKE ? OR LOWER(example) LIKE ?)"
            )

        sql = base_sql + level_clause + filter_sql + " ORDER BY RANDOM() LIMIT 50"
        if q_raw:
            params.extend([q, q, q, q, q])
        cur.execute(sql, tuple(params))
        rows = cur.fetchall()

        return [dict(r) for r in rows]
    except Exception as e:
        print(f"Error in search_words: {e}")
        return []
    finally:
        conn.close()


def get_words_by_subject_level(subject: str, level: str) -> List[Dict[str, Any]]:
    """Get words by subject, course language, and level from the database"""
    vocab_lang = vocab_language_for_subject(subject)
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(
        'SELECT * FROM words WHERE subject=? AND language=? AND level=? ORDER BY RANDOM() LIMIT 50',
        (subject, vocab_lang, level),
    )
    rows = cur.fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_words_by_subject(subject: str) -> List[Dict[str, Any]]:
    vocab_lang = vocab_language_for_subject(subject)
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(
        'SELECT * FROM words WHERE subject=? AND language=? ORDER BY RANDOM() LIMIT 100',
        (subject, vocab_lang),
    )
    rows = cur.fetchall()
    conn.close()
    return [dict(r) for r in rows]


def save_student_preference(user_id: int, preferred_translation: str):
    conn = get_conn()
    cur = conn.cursor()
    # Upsert style: delete existing then insert
    cur.execute('DELETE FROM student_preferences WHERE user_id=?', (user_id,))
    cur.execute('INSERT INTO student_preferences (user_id, preferred_translation) VALUES (?,?)', (user_id, preferred_translation))
    conn.commit()
    conn.close()


def get_student_preference(user_id: int) -> str:
    conn = get_conn()
    cur = conn.cursor()
    cur.execute('SELECT preferred_translation FROM student_preferences WHERE user_id=?', (user_id,))
    row = cur.fetchone()
    conn.close()
    return row['preferred_translation'] if row else None


def _sanitize_excluded_ids(exclude_word_ids: set[int] | None) -> list[int]:
    out: set[int] = set()
    for raw in (exclude_word_ids or set()):
        try:
            val = int(raw)
        except Exception:
            continue
        if val > 0:
            out.add(val)
    return sorted(out)


def _get_random_distractors(
    cur,
    subject,
    language,
    column,
    correct_value,
    count=3,
    exclude_word_ids: set[int] | None = None,
):
    # PostgreSQL: `SELECT DISTINCT ... ORDER BY RANDOM()` can fail when the ORDER BY
    # expression isn't included in the select list. We avoid DISTINCT here and
    # de-duplicate in Python (duplicates are still acceptable).
    sql = (
        f'SELECT {column} FROM words '
        f'WHERE subject=? AND language=? '
        f'AND {column} IS NOT NULL AND LOWER({column})!=? '
    )
    params: list[Any] = [subject, language, (correct_value or "").lower()]
    excluded = _sanitize_excluded_ids(exclude_word_ids)
    if excluded:
        placeholders = ",".join(["?"] * len(excluded))
        sql += f" AND id NOT IN ({placeholders}) "
        params.extend(excluded)
    sql += " ORDER BY RANDOM() LIMIT ?"
    params.append(int(count))
    cur.execute(sql, tuple(params))
    rows = cur.fetchall()
    out = []
    for r in rows:
        try:
            val = r.get(column)  # psycopg dict_row
        except Exception:
            try:
                val = r[0]  # sqlite row / tuple
            except Exception:
                val = None
        if val:
            out.append(val)
    return out

def _ensure_four_options(
    cur,
    subject: str,
    language: str,
    column: str,
    options: List[str],
    exclude_word_ids: set[int] | None = None,
) -> List[str]:
    """Pad options to 4 unique items (best effort) using random values from DB."""
    seen = []
    for o in options:
        if o and o not in seen:
            seen.append(o)
    # pull extra candidates until we reach 4 or exhaust tries
    tries = 0
    while len(seen) < 4 and tries < 5:
        need = 4 - len(seen)
        sql = (
            f"SELECT {column} FROM words "
            f"WHERE subject=? AND language=? AND {column} IS NOT NULL "
        )
        params: list[Any] = [subject, language]
        excluded = _sanitize_excluded_ids(exclude_word_ids)
        if excluded:
            placeholders = ",".join(["?"] * len(excluded))
            sql += f" AND id NOT IN ({placeholders}) "
            params.extend(excluded)
        sql += " ORDER BY RANDOM() LIMIT ?"
        params.append(need)
        cur.execute(sql, tuple(params))
        for r in cur.fetchall():
            try:
                val = r.get(column)  # psycopg dict_row
            except Exception:
                try:
                    val = r[0]
                except Exception:
                    val = None
            if val and val not in seen:
                seen.append(val)
        tries += 1
    # Final fallback: ensure we always have 4 options for Telegram polls.
    # Duplicates are acceptable; it is better than skipping the whole question.
    if not seen:
        fallback = "—"
    else:
        fallback = seen[0]
    while len(seen) < 4:
        seen.append(fallback)
    return seen[:4]


VOCAB_QUIZ_TYPES = (
    "multiple_choice",
    "gap_filling",
    "definition",
)


def vocab_instruction_for_question(question: Dict[str, Any], subject: str = "") -> str:
    """Human-readable per-question instruction based on the actual vocabulary question type."""
    qtype = str(question.get("question_type") or question.get("type") or "").strip().lower()
    ask = str(question.get("ask") or "").strip().lower()
    is_ru_subject = str(subject or "").strip().lower() == "russian"
    if is_ru_subject:
        if qtype == "gap_filling":
            return "Выберите подходящее слово для пропуска."
        if qtype == "definition":
            return "Выберите определение, которое подходит к слову."
        if qtype == "translation":
            return "Выберите правильный перевод слова."
        if qtype == "multiple_choice" and ask == "word":
            return "Выберите слово, которое подходит к переводу."
        if qtype == "multiple_choice" and ask == "translation":
            return "Выберите правильный перевод слова."
        return "Выберите правильный ответ."
    if qtype == "gap_filling":
        return "Choose the word that fits the blank."
    if qtype == "definition":
        return "Choose the definition that matches the word."
    if qtype == "translation":
        return "Choose the correct translation of the word."
    if qtype == "multiple_choice" and ask == "word":
        return "Choose the word that matches the translation."
    if qtype == "multiple_choice" and ask == "translation":
        return "Choose the correct translation of the word."
    return "Choose the correct answer."


def attach_vocab_instruction(question: Dict[str, Any], subject: str = "") -> Dict[str, Any]:
    instruction = vocab_instruction_for_question(question, subject)
    prompt = str(question.get("prompt") or "").strip()
    if prompt and not prompt.lower().startswith(instruction.lower()[:12]):
        question["prompt"] = f"{instruction}\n{prompt}"
    else:
        question["prompt"] = prompt or instruction
    question["instruction"] = instruction
    return question


def _fake_word_variants(word: str, count: int = 3, subject: str = "") -> list[str]:
    is_ru_subject = str(subject or "").strip().lower() == "russian"
    if is_ru_subject:
        base = re.sub(r"[^А-Яа-яЁё]", "", str(word or "").lower()) or "слово"
        seeds = [
            f"{base}о",
            f"{base[:-1]}ый" if len(base) > 3 else f"{base}ик",
            f"{base[:2]}{base[-1:]}{base[2:-1]}а" if len(base) > 4 else f"{base}ка",
            f"{base}ность",
            f"{base}ование",
        ]
    else:
        base = re.sub(r"[^A-Za-z]", "", str(word or "").lower()) or "word"
        seeds = [
            f"{base}y",
            f"{base[:-1]}ox" if len(base) > 3 else f"{base}ox",
            f"{base[:2]}{base[-1:]}{base[2:-1]}e" if len(base) > 4 else f"{base}en",
            f"{base}fuly",
            f"{base}menty",
        ]
    variants: list[str] = []
    for item in seeds:
        cleaned = item.strip().lower()
        if cleaned and cleaned != base and cleaned not in variants:
            variants.append(cleaned)
        if len(variants) >= count:
            break
    while len(variants) < count:
        variants.append(f"{base}{len(variants) + 2}x")
    return variants[:count]


def _safe_option_texts(values: list[Any], limit: int = 4) -> list[str]:
    out: list[str] = []
    seen: set[str] = set()
    for value in values:
        text = str(value or "").strip()
        if not text:
            continue
        key = text.lower()
        if key in seen:
            continue
        seen.add(key)
        out.append(text)
        if len(out) >= limit:
            break
    return out


def balanced_vocab_question_types(count: int) -> list[str]:
    """Return a shuffled, balanced vocabulary question-type plan."""
    total = max(1, int(count or 1))
    buckets: list[str] = []
    while len(buckets) < total:
        cycle = list(VOCAB_QUIZ_TYPES)
        random.shuffle(cycle)
        buckets.extend(cycle)
    return buckets[:total]


def generate_quiz(
    user_id: int,
    subject: str,
    level: str,
    count: int,
    qtype: str,
    preferred_translation: str,
    exclude_word_ids: set[int] | None = None,
    connection: Any | None = None,
) -> List[Dict[str, Any]]:
    """
    qtype: 'multiple_choice', 'gap_filling', 'definition'
    preferred_translation: 'uz' or 'ru' — which translation language to use when presenting translations
    Returns list of questions with options and correct answer index/value
    """
    # Mixed quizzes call this once per type/level. Reusing one connection
    # avoids paying the database connection setup cost 6–18 times while the
    # student is waiting on the "test is preparing" screen.
    owns_connection = connection is None
    conn = connection or get_conn()
    cur = conn.cursor()
    vocab_lang = vocab_language_for_subject(subject)
    excluded_ids = _sanitize_excluded_ids(exclude_word_ids)
    params: list[Any] = [subject, vocab_lang]
    sql = 'SELECT * FROM words WHERE subject=? AND language=?'
    if level:
        sql += ' AND level=?'
        params.append(level)
    if excluded_ids:
        placeholders = ",".join(["?"] * len(excluded_ids))
        sql += f" AND id NOT IN ({placeholders})"
        params.extend(excluded_ids)
    sql += ' ORDER BY RANDOM() LIMIT ?'
    params.append(count)
    cur.execute(sql, tuple(params))
    rows = cur.fetchall()
    words = [dict(r) for r in rows]

    questions = []
    for w in words:
        word_id = int(w.get('id') or 0)
        if qtype == 'translation':
            # Tarjimasini topish: so'z ko'rsatiladi, to'g'ri tarjima tanlanadi
            subj = (subject or '').lower()
            if subj == 'russian':
                # Rus kursi: ruscha so'z → o'zbekcha tarjima
                pref_col = 'translation_uz'
            else:
                pref_col = 'translation_uz' if preferred_translation == 'uz' else 'translation_ru'
            correct = (w.get(pref_col) or '').strip()
            if not correct:
                correct = (w.get('translation_uz') or w.get('translation_ru') or '').strip()
            if not correct:
                continue
            distractors = _get_random_distractors(
                cur,
                subject,
                w['language'],
                pref_col,
                correct,
                3,
                exclude_word_ids=excluded_ids,
            )
            options = _ensure_four_options(
                cur,
                subject,
                w['language'],
                pref_col,
                [correct] + distractors,
                exclude_word_ids=excluded_ids,
            )
            random.shuffle(options)
            questions.append(attach_vocab_instruction({
                'type': 'translation',
                'prompt': w.get('word') or '',
                'options': options,
                'correct': correct,
                'word_id': word_id,
                'question_type': qtype,
            }, subject))

        elif qtype == 'multiple_choice':
            # Decide whether to ask for translation or the word itself randomly
            ask_translation = random.choice([True, False])
            if ask_translation:
                correct = w['translation_uz'] if preferred_translation == 'uz' else w['translation_ru']
                if not correct:
                    # fallback to other
                    correct = w['translation_uz'] or w['translation_ru'] or w['word']
                distractors = _get_random_distractors(
                    cur,
                    subject,
                    w['language'],
                    'translation_uz' if preferred_translation == 'uz' else 'translation_ru',
                    correct,
                    3,
                    exclude_word_ids=excluded_ids,
                )
                options = _ensure_four_options(
                    cur,
                    subject,
                    w['language'],
                    'translation_uz' if preferred_translation == 'uz' else 'translation_ru',
                    [correct] + distractors,
                    exclude_word_ids=excluded_ids,
                )
                random.shuffle(options)
                questions.append(
                    attach_vocab_instruction({
                        'type': 'multiple_choice',
                        'prompt': w['word'],
                        'options': options,
                        'correct': correct,
                        'ask': 'translation',
                        'word_id': word_id,
                        'question_type': qtype,
                    }, subject)
                )
            else:
                # Ask to choose the correct word given a translation
                correct = w['word']
                col = 'translation_uz' if preferred_translation == 'uz' else 'translation_ru'
                prompt = w[col] or w['translation_uz'] or w['translation_ru'] or ''
                distractors = _get_random_distractors(
                    cur,
                    subject,
                    w['language'],
                    'word',
                    correct,
                    3,
                    exclude_word_ids=excluded_ids,
                )
                options = _ensure_four_options(
                    cur,
                    subject,
                    w['language'],
                    'word',
                    [correct] + distractors,
                    exclude_word_ids=excluded_ids,
                )
                random.shuffle(options)
                questions.append(
                    attach_vocab_instruction({
                        'type': 'multiple_choice',
                        'prompt': prompt,
                        'options': options,
                        'correct': correct,
                        'ask': 'word',
                        'word_id': word_id,
                        'question_type': qtype,
                    }, subject)
                )

        elif qtype == 'gap_filling':
            example = w.get('example') or ''
            if not example or w['word'] == '':
                # fallback to multiple choice
                distractors = _get_random_distractors(
                    cur,
                    subject,
                    w['language'],
                    'word',
                    w['word'],
                    3,
                    exclude_word_ids=excluded_ids,
                )
                options = _ensure_four_options(
                    cur,
                    subject,
                    w['language'],
                    'word',
                    [w['word']] + distractors,
                    exclude_word_ids=excluded_ids,
                )
                random.shuffle(options)
                questions.append(attach_vocab_instruction({
                    'type': 'multiple_choice',
                    'prompt': w['word'],
                    'options': options,
                    'correct': w['word'],
                    'ask': 'word',
                    'word_id': word_id,
                    'question_type': qtype,
                }, subject))
                continue
            # replace first occurrence of the word (case-insensitive)
            import re
            pattern = re.compile(re.escape(w['word']), re.IGNORECASE)
            blanked = pattern.sub('_____', example, count=1)
            # options: correct word + distractors
            distractors = _get_random_distractors(
                cur,
                subject,
                w['language'],
                'word',
                w['word'],
                3,
                exclude_word_ids=excluded_ids,
            )
            options = _ensure_four_options(
                cur,
                subject,
                w['language'],
                'word',
                [w['word']] + distractors,
                exclude_word_ids=excluded_ids,
            )
            random.shuffle(options)
            questions.append(
                attach_vocab_instruction({
                    'type': 'gap_filling',
                    'prompt': blanked,
                    'options': options,
                    'correct': w['word'],
                    'word_id': word_id,
                    'question_type': qtype,
                }, subject)
            )

        elif qtype == 'definition':
            correct = w.get('definition') or ''
            if not correct:
                # fallback to multiple choice on word
                distractors = _get_random_distractors(
                    cur,
                    subject,
                    w['language'],
                    'word',
                    w['word'],
                    3,
                    exclude_word_ids=excluded_ids,
                )
                options = _ensure_four_options(
                    cur,
                    subject,
                    w['language'],
                    'word',
                    [w['word']] + distractors,
                    exclude_word_ids=excluded_ids,
                )
                random.shuffle(options)
                questions.append(attach_vocab_instruction({
                    'type': 'multiple_choice',
                    'prompt': w['word'],
                    'options': options,
                    'correct': w['word'],
                    'ask': 'word',
                    'word_id': word_id,
                    'question_type': qtype,
                }, subject))
                continue
            distractors = _get_random_distractors(
                cur,
                subject,
                w['language'],
                'definition',
                correct,
                3,
                exclude_word_ids=excluded_ids,
            )
            options = _ensure_four_options(
                cur,
                subject,
                w['language'],
                'definition',
                [correct] + distractors,
                exclude_word_ids=excluded_ids,
            )
            random.shuffle(options)
            questions.append(
                attach_vocab_instruction({
                    'type': 'definition',
                    'prompt': w['word'],
                    'options': options,
                    'correct': correct,
                    'word_id': word_id,
                    'question_type': qtype,
                }, subject)
            )

    if owns_connection:
        conn.close()
    return questions


def generate_balanced_mixed_quiz(
    user_id: int,
    subject: str,
    levels: list[str] | tuple[str, ...] | None,
    count: int,
    preferred_translation: str,
    cooldown_word_ids_by_type: dict[str, set[int]] | None = None,
) -> List[Dict[str, Any]]:
    """Generate a fixed-size mixed vocabulary quiz with balanced question types."""
    target_count = max(1, int(count or 1))
    level_list = [str(level or "").strip().upper() for level in (levels or []) if str(level or "").strip()]
    if not level_list:
        level_list = [""]
    plan = balanced_vocab_question_types(target_count)
    budgets = {qtype: plan.count(qtype) for qtype in VOCAB_QUIZ_TYPES}
    by_type: dict[str, list[dict[str, Any]]] = {qtype: [] for qtype in VOCAB_QUIZ_TYPES}
    used_word_ids: set[int] = set()
    cooldowns = cooldown_word_ids_by_type or {}
    shared_connection = get_conn()

    def collect_for(qtype: str, *, use_cooldown: bool) -> None:
        need_total = int(budgets.get(qtype) or 0)
        if need_total <= len(by_type[qtype]):
            return
        excluded_word_ids = set(used_word_ids)
        if use_cooldown:
            excluded_word_ids.update(cooldowns.get(qtype) or set())
        for level in level_list:
            need = need_total - len(by_type[qtype])
            if need <= 0:
                break
            questions = generate_quiz(
                user_id=user_id,
                subject=subject,
                level=level,
                count=need,
                qtype=qtype,
                preferred_translation=preferred_translation,
                exclude_word_ids=excluded_word_ids,
                connection=shared_connection,
            )
            for question in questions:
                wid = int(question.get("word_id") or 0)
                if wid > 0 and wid in used_word_ids:
                    continue
                question["question_type"] = qtype
                by_type[qtype].append(question)
                if wid > 0:
                    used_word_ids.add(wid)
                    excluded_word_ids.add(wid)
                if len(by_type[qtype]) >= need_total:
                    break

    try:
        for qtype in VOCAB_QUIZ_TYPES:
            collect_for(qtype, use_cooldown=True)
        if sum(len(items) for items in by_type.values()) < target_count:
            for qtype in VOCAB_QUIZ_TYPES:
                collect_for(qtype, use_cooldown=False)

        mixed: list[dict[str, Any]] = []
        for qtype in plan:
            if by_type.get(qtype):
                mixed.append(by_type[qtype].pop(0))
            else:
                for fallback_type in VOCAB_QUIZ_TYPES:
                    if by_type.get(fallback_type):
                        mixed.append(by_type[fallback_type].pop(0))
                        break
            if len(mixed) >= target_count:
                break

        leftovers = [question for qtype in VOCAB_QUIZ_TYPES for question in by_type.get(qtype, [])]
        random.shuffle(leftovers)
        for question in leftovers:
            if len(mixed) >= target_count:
                break
            mixed.append(question)
        return mixed[:target_count]
    finally:
        shared_connection.close()


def export_words_to_xlsx(subject: str) -> Tuple[io.BytesIO, str]:
    conn = get_conn()
    cur = conn.cursor()
    cur.execute('SELECT * FROM words WHERE subject=? ORDER BY word', (subject,))
    rows = cur.fetchall()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = f"{subject} Vocabulary"
    
    # Use the same headers as import format
    headers = ['word', 'level', 'translation_uz', 'translation_ru', 'definition', 'example']
    ws.append(headers)
    
    for r in rows:
        row_dict = dict(r)
        ws.append([
            row_dict['word'], row_dict.get('level') or '', row_dict.get('translation_uz') or '', row_dict.get('translation_ru') or '',
            row_dict.get('definition') or '', row_dict.get('example') or ''
        ])

    # Auto-adjust column widths for better readability
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[column_letter].width = adjusted_width

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    
    # Use the same naming convention as import
    filename = f"Vocabulary_importing_List_for_{subject}.xlsx"
    return bio, filename


def export_subject_dcoin_history_to_xlsx(
    subject: str,
    owner_admin_id: int | None = None,
    lang: str = "uz",
) -> Tuple[io.BytesIO, str]:
    from db import get_subject_dcoin_history_rows
    from i18n import t

    _ = subject
    subj = "GLOBAL"
    rows = get_subject_dcoin_history_rows(subj, owner_admin_id=owner_admin_id)
    if not rows:
        raise ValueError("empty_subject_dcoin_history")

    wb = Workbook()
    ws = wb.active
    ws.title = "Dcoin History"
    headers = [
        "Sana/Vaqt",
        "Fan",
        "Login ID",
        "Ism",
        "Familya",
        "D'coin o'zgarish (+/-)",
        "D'points o'zgarish (+/-)",
        "Holat",
        "Sabab",
    ]
    ws.append(headers)

    earn_label = t(lang, "staff_dcoin_export_status_earn")
    lose_label = t(lang, "staff_dcoin_export_status_lose")
    for row in rows:
        change = float(row.get("dcoin_change") or 0)
        status_label = earn_label if change >= 0 else lose_label
        ws.append(
            [
                row.get("created_at") or "",
                row.get("subject") or "GLOBAL",
                row.get("login_id") or "",
                row.get("first_name") or "",
                row.get("last_name") or "",
                f"{change:+.1f}",
                f"{float(row.get('dpoints_change') or 0):+.1f}",
                status_label,
                row.get("change_type") or "",
            ]
        )

    for column in ws.columns:
        max_length = 0
        col_letter = column[0].column_letter
        for cell in column:
            cell_value = "" if cell.value is None else str(cell.value)
            if len(cell_value) > max_length:
                max_length = len(cell_value)
        ws.column_dimensions[col_letter].width = min(max_length + 2, 40)

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    safe_subject = "GLOBAL"
    return bio, f"dcoin_history_{safe_subject}.xlsx"


def export_all_groups_to_xlsx() -> Tuple[io.BytesIO, str]:
    from db import get_all_groups, get_group_users
    groups = get_all_groups()
    wb = Workbook()
    ws = wb.active
    ws.append(['group_id', 'group_name', 'level', 'teacher_id', 'teacher_name', 'student_id', 'student_login', 'student_name', 'student_subject', 'student_level'])
    for g in groups:
        users = get_group_users(g['id'])
        teacher_name = ''
        if g.get('teacher_id'):
            from db import get_user_by_id
            t = get_user_by_id(g['teacher_id'])
            if t:
                teacher_name = f"{t.get('first_name','')} {t.get('last_name','')}"
        if not users:
            ws.append([g['id'], g['name'], g.get('level') or '', g.get('teacher_id'), teacher_name, '', '', '', '', ''])
        else:
            for u in users:
                ws.append([g['id'], g['name'], g.get('level') or '', g.get('teacher_id'), teacher_name, u['id'], u.get('login_id'), f"{u.get('first_name','')} {u.get('last_name','')}", u.get('subject'), u.get('level')])

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio, 'groups_with_students.xlsx'


def export_groups_for_teacher_to_xlsx(teacher_id: int) -> Tuple[io.BytesIO, str]:
    from db import get_groups_by_teacher, get_group_users, get_user_by_id
    groups = get_groups_by_teacher(teacher_id)
    wb = Workbook()
    ws = wb.active
    ws.append(['group_id', 'group_name', 'level', 'teacher_id', 'teacher_name', 'student_id', 'student_login', 'student_name', 'student_subject', 'student_level'])
    for g in groups:
        users = get_group_users(g['id'])
        teacher_name = ''
        if g.get('teacher_id'):
            t = get_user_by_id(g['teacher_id'])
            if t:
                teacher_name = f"{t.get('first_name','')} {t.get('last_name','')}"
        if not users:
            ws.append([g['id'], g['name'], g.get('level') or '', g.get('teacher_id'), teacher_name, '', '', '', '', ''])
        else:
            for u in users:
                ws.append([g['id'], g['name'], g.get('level') or '', g.get('teacher_id'), teacher_name, u['id'], u.get('login_id'), f"{u.get('first_name','')} {u.get('last_name','')}", u.get('subject'), u.get('level')])

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio, f'teacher_{teacher_id}_groups.xlsx'


def export_students_to_xlsx() -> Tuple[io.BytesIO, str]:
    from db import get_all_users
    users = get_all_users()
    wb = Workbook()
    ws = wb.active
    ws.append(['id', 'login_id', 'first_name', 'last_name', 'phone', 'subject', 'level', 'telegram_id', 'group_id', 'access_enabled'])
    for u in users:
        ws.append([u.get('id'), u.get('login_id'), u.get('first_name'), u.get('last_name'), u.get('phone'), u.get('subject'), u.get('level'), u.get('telegram_id'), u.get('group_id'), u.get('access_enabled')])

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio, 'all_students.xlsx'


def export_full_system_to_xlsx() -> Tuple[io.BytesIO, str]:
    """
    Export all public PostgreSQL tables into a single XLSX workbook.
    Each table becomes a separate worksheet.
    """
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        SELECT table_name
        FROM information_schema.tables
        WHERE table_schema = 'public' AND table_type = 'BASE TABLE'
        ORDER BY table_name
    """)
    tables = [r["table_name"] for r in cur.fetchall()]

    wb = Workbook()
    # Replace default sheet with a small info sheet
    ws_info = wb.active
    ws_info.title = 'meta'
    from datetime import datetime
    ws_info.append(['exported_at_utc', datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')])
    ws_info.append(['tables', ', '.join(tables)])

    for table in tables:
        # Excel sheet title max length is 31
        sheet_name = table[:31]
        ws = wb.create_sheet(title=sheet_name)
        cur.execute(f'SELECT * FROM "{table}"')
        rows = cur.fetchall()
        headers = [d[0] for d in (cur.description or [])]
        if headers:
            ws.append(headers)
        for r in rows:
            ws.append([r[h] for h in headers])

    conn.close()
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio, 'diamond_full_export.xlsx'


def export_all_attendance_to_xlsx(owner_admin_id: int | None = None) -> Tuple[io.BytesIO, str]:
    """Export attendance to Excel. If owner_admin_id given, only that admin's groups."""
    from db import get_conn, get_group
    
    conn = get_conn()
    cur = conn.cursor()
    
    if owner_admin_id is not None:
        cur.execute("""
            SELECT a.user_id, a.group_id, a.date, a.status, 
                   u.first_name, u.last_name, u.login_id, u.telegram_id,
                   g.name as group_name, g.level as group_level
            FROM attendance a
            JOIN users u ON a.user_id = u.id
            JOIN groups g ON a.group_id = g.id
            WHERE (
                g.owner_admin_id = ?
                OR u.id IN (
                    SELECT student_id FROM admin_student_shares
                    WHERE peer_admin_id = ? AND status = 'active'
                )
            )
            ORDER BY a.date DESC, g.name, u.first_name
        """, (owner_admin_id, owner_admin_id))
    else:
        cur.execute("""
            SELECT a.user_id, a.group_id, a.date, a.status, 
                   u.first_name, u.last_name, u.login_id, u.telegram_id,
                   g.name as group_name, g.level as group_level
            FROM attendance a
            JOIN users u ON a.user_id = u.id
            JOIN groups g ON a.group_id = g.id
            ORDER BY a.date DESC, g.name, u.first_name
        """)
    
    attendance_data = cur.fetchall()
    conn.close()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"
    
    # Headers
    headers = ['Date', 'Group', 'Group Level', 'Student ID', 'First Name', 'Last Name', 'Telegram ID', 'Status']
    ws.append(headers)
    
    # Data
    for row in attendance_data:
        ws.append([
            row['date'],
            row['group_name'],
            row['group_level'],
            row['login_id'],
            row['first_name'],
            row['last_name'],
            row['telegram_id'] or '-',
            row['status']
        ])
    
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio, 'attendance_all_export.xlsx'


def export_group_attendance_to_xlsx(group_id: int) -> Tuple[io.BytesIO, str]:
    """Export attendance data for a specific group to Excel"""
    from db import get_conn, get_group
    
    group = get_group(group_id)
    if not group:
        raise ValueError("Group not found")
    
    conn = get_conn()
    cur = conn.cursor()
    
    # Get attendance for this group
    cur.execute("""
        SELECT a.user_id, a.date, a.status, 
               u.first_name, u.last_name, u.login_id, u.telegram_id
        FROM attendance a
        JOIN users u ON a.user_id = u.id
        WHERE a.group_id = ?
        ORDER BY a.date DESC, u.first_name
    """, (group_id,))
    
    attendance_data = cur.fetchall()
    conn.close()
    
    wb = Workbook()
    ws = wb.active
    ws.title = f"Attendance - {group['name']}"
    
    # Headers
    headers = ['Date', 'Student ID', 'First Name', 'Last Name', 'Telegram ID', 'Status']
    ws.append(headers)
    
    # Data
    for row in attendance_data:
        ws.append([
            row['date'],
            row['login_id'],
            row['first_name'],
            row['last_name'],
            row['telegram_id'] or '-',
            row['status']
        ])
    
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio, f'attendance_{group["name"]}_export.xlsx'
